"""API routes for SEO Crawler."""
import io
import asyncio
from collections import defaultdict
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.core.database import get_db
from app.models.models import Project, Crawl, Page
from app.schemas.schemas import (
    ProjectCreate, ProjectResponse,
    CrawlCreate, CrawlResponse,
    PageSummary, PageDetail, PageTableRow, CrawlSummary,
    DuplicateGroup, StatusCodeGroup, IssueGroup,
)
from app.crawler.engine import CrawlEngine, active_crawls

router = APIRouter()


# ─── Projects ───────────────────────────────────────────────
@router.post("/projects", response_model=ProjectResponse)
async def create_project(data: ProjectCreate, db: AsyncSession = Depends(get_db)):
    project = Project(name=data.name, url=data.url)
    db.add(project)
    await db.commit()
    await db.refresh(project)
    return project


@router.get("/projects", response_model=list[ProjectResponse])
async def list_projects(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Project).order_by(Project.created_at.desc()))
    return result.scalars().all()


@router.get("/projects/{project_id}", response_model=ProjectResponse)
async def get_project(project_id: int, db: AsyncSession = Depends(get_db)):
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@router.delete("/projects/{project_id}")
async def delete_project(project_id: int, db: AsyncSession = Depends(get_db)):
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    await db.delete(project)
    await db.commit()
    return {"message": "Project deleted"}


# ─── Crawls ─────────────────────────────────────────────────
async def _run_crawl(crawl_id: int, base_url: str):
    """Background task to execute the crawl."""
    engine = CrawlEngine(crawl_id, base_url)
    await engine.run()


@router.post("/crawls", response_model=CrawlResponse)
async def start_crawl(data: CrawlCreate, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)):
    project = await db.get(Project, data.project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    crawl = Crawl(project_id=project.id, status="pending")
    db.add(crawl)
    await db.commit()
    await db.refresh(crawl)

    background_tasks.add_task(_run_crawl, crawl.id, project.url)
    return crawl


@router.get("/crawls/{crawl_id}", response_model=CrawlResponse)
async def get_crawl(crawl_id: int, db: AsyncSession = Depends(get_db)):
    crawl = await db.get(Crawl, crawl_id)
    if not crawl:
        raise HTTPException(status_code=404, detail="Crawl not found")
    return crawl


@router.get("/projects/{project_id}/crawls", response_model=list[CrawlResponse])
async def list_crawls(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Crawl).where(Crawl.project_id == project_id).order_by(Crawl.created_at.desc())
    )
    return result.scalars().all()


# ─── Crawl Control (Pause / Stop / Resume) ──────────────────
@router.post("/crawls/{crawl_id}/pause")
async def pause_crawl(crawl_id: int, db: AsyncSession = Depends(get_db)):
    crawl = await db.get(Crawl, crawl_id)
    if not crawl:
        raise HTTPException(status_code=404, detail="Crawl not found")
    if crawl.status != "running":
        raise HTTPException(status_code=400, detail=f"Cannot pause crawl with status '{crawl.status}'")

    engine = active_crawls.get(crawl_id)
    if not engine:
        raise HTTPException(status_code=400, detail="Crawl engine not found in memory")

    engine.pause()
    crawl.status = "paused"
    await db.commit()
    return {"message": "Crawl paused", "status": "paused"}


@router.post("/crawls/{crawl_id}/resume")
async def resume_crawl(crawl_id: int, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)):
    crawl = await db.get(Crawl, crawl_id)
    if not crawl:
        raise HTTPException(status_code=404, detail="Crawl not found")

    if crawl.status == "paused":
        # Resume in-memory engine
        engine = active_crawls.get(crawl_id)
        if not engine:
            raise HTTPException(status_code=400, detail="Crawl engine not found in memory")
        engine.resume()
        crawl.status = "running"
        await db.commit()
        return {"message": "Crawl resumed", "status": "running"}

    elif crawl.status == "stopped":
        # Re-start engine, loading already-crawled URLs from DB
        project = await db.get(Project, crawl.project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        async def _resume_crawl(cid: int, url: str):
            engine = CrawlEngine(cid, url)
            await engine.run(resume_from_stopped=True)

        background_tasks.add_task(_resume_crawl, crawl.id, project.url)
        return {"message": "Crawl resuming from stopped state", "status": "running"}

    else:
        raise HTTPException(status_code=400, detail=f"Cannot resume crawl with status '{crawl.status}'")


@router.post("/crawls/{crawl_id}/stop")
async def stop_crawl(crawl_id: int, db: AsyncSession = Depends(get_db)):
    crawl = await db.get(Crawl, crawl_id)
    if not crawl:
        raise HTTPException(status_code=404, detail="Crawl not found")
    if crawl.status not in ("running", "paused"):
        raise HTTPException(status_code=400, detail=f"Cannot stop crawl with status '{crawl.status}'")

    engine = active_crawls.get(crawl_id)
    if engine:
        engine.stop()
        # The engine will set status to "stopped" when workers finish
    else:
        # Engine not in memory (server restarted?), just update DB
        crawl.status = "stopped"
        crawl.completed_at = __import__("datetime").datetime.utcnow()
        await db.commit()

    return {"message": "Crawl stopping", "status": "stopped"}


# ─── Pages ──────────────────────────────────────────────────
@router.get("/crawls/{crawl_id}/pages", response_model=list[PageSummary])
async def list_pages(crawl_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Page).where(Page.crawl_id == crawl_id).order_by(Page.id.asc())
    )
    pages = result.scalars().all()
    return [
        PageSummary(
            id=p.id,
            url=p.url,
            status_code=p.status_code,
            title=p.title,
            score=p.score,
            issues_count=len(p.issues) if p.issues else 0,
            response_time=p.response_time,
        )
        for p in pages
    ]


@router.get("/crawls/{crawl_id}/pages/table", response_model=list[PageTableRow])
async def list_pages_table(crawl_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Page).where(Page.crawl_id == crawl_id).order_by(Page.id.asc())
    )
    pages = result.scalars().all()
    return [
        PageTableRow(
            id=p.id,
            url=p.url,
            status_code=p.status_code,
            score=p.score,
            title=p.title,
            title_length=p.title_length,
            meta_description_length=p.meta_description_length,
            canonical_url=p.canonical_url,
            is_noindex=p.is_noindex or False,
            is_nofollow_meta=p.is_nofollow_meta or False,
            h1_count=p.h1_count or 0,
            h2_count=p.h2_count or 0,
            total_images=p.total_images or 0,
            images_without_alt=p.images_without_alt or 0,
            images_with_empty_alt=p.images_with_empty_alt or 0,
            internal_links=p.internal_links or 0,
            external_links=p.external_links or 0,
            nofollow_links=p.nofollow_links or 0,
            word_count=p.word_count or 0,
            code_to_text_ratio=p.code_to_text_ratio,
            has_schema_markup=p.has_schema_markup or False,
            has_hreflang=p.has_hreflang or False,
            has_viewport_meta=p.has_viewport_meta or False,
            has_lazy_loading=p.has_lazy_loading or False,
            has_placeholders=p.has_placeholders or False,
            response_time=p.response_time,
            issues_count=len(p.issues) if p.issues else 0,
        )
        for p in pages
    ]


@router.get("/pages/{page_id}", response_model=PageDetail)
async def get_page(page_id: int, db: AsyncSession = Depends(get_db)):
    page = await db.get(Page, page_id)
    if not page:
        raise HTTPException(status_code=404, detail="Page not found")
    return page


# ─── Dashboard / Summary ────────────────────────────────────
@router.get("/crawls/{crawl_id}/summary")
async def get_crawl_summary(crawl_id: int, db: AsyncSession = Depends(get_db)):
    crawl = await db.get(Crawl, crawl_id)
    if not crawl:
        raise HTTPException(status_code=404, detail="Crawl not found")

    result = await db.execute(select(Page).where(Page.crawl_id == crawl_id))
    pages = result.scalars().all()

    if not pages:
        raise HTTPException(status_code=404, detail="No pages found for this crawl")

    total = len(pages)

    # Separate redirect pages — they should NOT be counted for content/SEO issues
    REDIRECT_CODES = {301, 302, 303, 307, 308}
    content_pages = [p for p in pages if (p.status_code or 0) >= 200 and (p.status_code or 0) < 300]
    content_total = len(content_pages) or 1  # avoid division by zero

    avg_score = sum(p.score or 0 for p in content_pages) / content_total

    # Count issues by severity (only non-redirect pages for content issues)
    critical = 0
    warnings = 0
    info_count = 0
    issue_map = defaultdict(list)  # type -> [{url, page_id, detail}]

    for p in content_pages:
        for i in (p.issues or []):
            sev = i.get("severity", "")
            if sev == "critical":
                critical += 1
            elif sev == "warning":
                warnings += 1
            elif sev == "info":
                info_count += 1
            itype = i.get("type", "unknown")
            issue_map[itype].append({
                "url": p.url, "page_id": p.id,
                "detail": i.get("message", ""),
            })
    # Also count redirect issues separately
    for p in pages:
        if (p.status_code or 0) in REDIRECT_CODES:
            for i in (p.issues or []):
                if i.get("type") == "redirect":
                    warnings += 1
                    issue_map["redirect"].append({
                        "url": p.url, "page_id": p.id,
                        "detail": i.get("message", ""),
                    })

    # --- Duplicate titles (exclude redirects) ---
    title_groups = defaultdict(list)
    for p in content_pages:
        if p.title:
            title_groups[p.title].append({"url": p.url, "page_id": p.id})
    duplicate_titles = [
        DuplicateGroup(value=title, pages=pg, count=len(pg))
        for title, pg in title_groups.items() if len(pg) > 1
    ]

    # --- Duplicate meta descriptions (exclude redirects) ---
    meta_groups = defaultdict(list)
    for p in content_pages:
        if p.meta_description:
            meta_groups[p.meta_description].append({"url": p.url, "page_id": p.id})
    duplicate_metas = [
        DuplicateGroup(value=desc, pages=pg, count=len(pg))
        for desc, pg in meta_groups.items() if len(pg) > 1
    ]

    # --- Status code breakdown (ALL pages including redirects) ---
    status_groups = defaultdict(list)
    for p in pages:
        if p.status_code:
            status_groups[p.status_code].append({"url": p.url, "page_id": p.id})
    status_breakdown = [
        StatusCodeGroup(status_code=code, count=len(pg), pages=pg)
        for code, pg in sorted(status_groups.items())
    ]

    # --- Canonical issues (exclude redirects) ---
    canonical_issues = []
    for p in content_pages:
        if p.canonical_issues and len(p.canonical_issues) > 0:
            canonical_issues.append({
                "url": p.url, "page_id": p.id,
                "canonical_url": p.canonical_url,
                "issues": p.canonical_issues,
            })

    # --- Noindex / Nofollow (exclude redirects) ---
    noindex_pages = [{"url": p.url, "page_id": p.id} for p in content_pages if p.is_noindex]
    nofollow_pages = [{"url": p.url, "page_id": p.id, "nofollow_internal": p.nofollow_internal_links} for p in content_pages if p.is_nofollow_meta]

    # --- Images missing alt (exclude redirects) ---
    pages_missing_alt = []
    total_images_missing = 0
    pages_empty_alt = []
    total_images_empty_alt = 0
    for p in content_pages:
        if p.images_without_alt and p.images_without_alt > 0:
            sample_img = (p.images_without_alt_urls or [None])[0]
            pages_missing_alt.append({
                "url": p.url, "page_id": p.id,
                "missing_count": p.images_without_alt,
                "total_images": p.total_images,
                "sample_image_url": sample_img,
            })
            total_images_missing += p.images_without_alt
        if p.images_with_empty_alt and p.images_with_empty_alt > 0:
            sample_img = (p.images_with_empty_alt_urls or [None])[0] if hasattr(p, 'images_with_empty_alt_urls') and p.images_with_empty_alt_urls else None
            pages_empty_alt.append({
                "url": p.url, "page_id": p.id,
                "empty_count": p.images_with_empty_alt,
                "total_images": p.total_images,
                "sample_image_url": sample_img,
            })
            total_images_empty_alt += p.images_with_empty_alt

    # --- Hreflang issues (exclude redirects) ---
    hreflang_issues = []
    for p in content_pages:
        if p.hreflang_issues and len(p.hreflang_issues) > 0:
            hreflang_issues.append({
                "url": p.url, "page_id": p.id,
                "issues": p.hreflang_issues,
                "entries": p.hreflang_entries,
            })

    # --- Content issues (exclude redirects) ---
    thin_content = [{"url": p.url, "page_id": p.id, "word_count": p.word_count} for p in content_pages if p.word_count and p.word_count < 300]
    low_ratio = [{"url": p.url, "page_id": p.id, "ratio": p.code_to_text_ratio} for p in content_pages if p.code_to_text_ratio is not None and p.code_to_text_ratio < 10]
    placeholder_pages = [{"url": p.url, "page_id": p.id, "content": p.placeholder_content} for p in content_pages if p.has_placeholders]

    # --- Structure (exclude redirects and errors) ---
    missing_title = sum(1 for p in content_pages if not p.title)
    missing_meta = sum(1 for p in content_pages if not p.meta_description)
    missing_h1 = sum(1 for p in content_pages if p.h1_count == 0)
    missing_viewport = sum(1 for p in content_pages if not p.has_viewport_meta)
    short_title = sum(1 for p in content_pages if p.title_length and 0 < p.title_length < 30)
    long_title = sum(1 for p in content_pages if p.title_length and p.title_length > 60)
    short_meta_desc = sum(1 for p in content_pages if p.meta_description_length and 0 < p.meta_description_length < 120)
    long_meta_desc = sum(1 for p in content_pages if p.meta_description_length and p.meta_description_length > 160)
    multiple_h1 = sum(1 for p in content_pages if p.h1_count and p.h1_count > 1)
    missing_og_title = sum(1 for p in content_pages if not p.og_title)
    missing_og_image = sum(1 for p in content_pages if not p.og_image)
    no_lazy = sum(1 for p in content_pages if not p.has_lazy_loading)

    # --- Performance (all pages) ---
    avg_resp = sum(p.response_time or 0 for p in pages) / total
    slow_pages = [{"url": p.url, "page_id": p.id, "response_time": p.response_time} for p in pages if p.response_time and p.response_time > 3]

    # --- Schema (exclude redirects) ---
    no_schema = sum(1 for p in content_pages if not p.has_schema_markup)
    missing_canonical = sum(1 for p in content_pages if p.canonical_issues and "missing" in p.canonical_issues)

    # --- Issue groups for the grouped issues table ---
    issue_groups = []
    severity_map = {
        "missing_title": "critical", "missing_meta_description": "critical",
        "missing_h1": "critical", "missing_viewport": "critical",
        "placeholder_content": "critical", "http_error": "critical",
        "noindex": "warning", "nofollow_meta": "warning",
        "short_title": "warning", "long_title": "warning",
        "short_meta_description": "warning", "long_meta_description": "warning",
        "missing_canonical": "warning", "canonical_external": "warning",
        "canonical_mismatch": "warning", "canonical_relative": "warning",
        "images_missing_alt": "warning", "images_empty_alt": "warning",
        "thin_content": "warning", "low_text_ratio": "warning",
        "multiple_h1": "warning", "nofollow_internal": "warning",
        "hreflang_issue": "warning", "slow_response": "warning",
        "redirect": "info",
        "no_schema_markup": "info", "no_lazy_loading": "info",
        "missing_og_title": "info", "missing_og_description": "info",
        "missing_og_image": "info", "high_text_ratio": "info",
    }
    for itype, pages_list in issue_map.items():
        issue_groups.append(IssueGroup(
            category=itype,
            severity=severity_map.get(itype, "info"),
            count=len(pages_list),
            pages=pages_list[:50],  # cap at 50 per group
        ))
    issue_groups.sort(key=lambda g: ({"critical": 0, "warning": 1, "info": 2}.get(g.severity, 3), -g.count))

    return {
        "total_pages": total,
        "avg_score": round(avg_score, 1),
        "critical_issues": critical,
        "warnings": warnings,
        "info_issues": info_count,
        "duplicate_titles": duplicate_titles,
        "duplicate_meta_descriptions": duplicate_metas,
        "status_code_breakdown": status_breakdown,
        "canonical_issues": canonical_issues,
        "noindex_pages": noindex_pages,
        "nofollow_pages": nofollow_pages,
        "pages_missing_alt": pages_missing_alt,
        "total_images_missing_alt": total_images_missing,
        "pages_empty_alt": pages_empty_alt,
        "total_images_empty_alt": total_images_empty_alt,
        "hreflang_issues": hreflang_issues,
        "thin_content_pages": thin_content,
        "low_text_ratio_pages": low_ratio,
        "placeholder_pages": placeholder_pages,
        "pages_missing_title": missing_title,
        "pages_missing_meta": missing_meta,
        "pages_missing_h1": missing_h1,
        "pages_missing_viewport": missing_viewport,
        "avg_response_time": round(avg_resp, 3),
        "slow_pages": slow_pages,
        "robots_txt_status": crawl.robots_txt_status,
        "sitemaps_found": crawl.sitemaps_found,
        "pages_without_schema": no_schema,
        "pages_missing_canonical": missing_canonical,
        "issue_groups": issue_groups,
    }


# ─── Excel Export ──────────────────────────────────────────
@router.get("/crawls/{crawl_id}/export/excel")
async def export_crawl_excel(crawl_id: int, db: AsyncSession = Depends(get_db)):
    """Generate an Excel report with separate sheets per issue type."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlalchemy.orm import selectinload

    crawl_result = await db.execute(
        select(Crawl).options(selectinload(Crawl.project)).where(Crawl.id == crawl_id)
    )
    crawl = crawl_result.scalar_one_or_none()
    if not crawl:
        raise HTTPException(status_code=404, detail="Crawl not found")

    result = await db.execute(select(Page).where(Page.crawl_id == crawl_id))
    pages = result.scalars().all()
    if not pages:
        raise HTTPException(status_code=404, detail="No pages found")

    REDIRECT_CODES = {301, 302, 303, 307, 308}
    content_pages = [p for p in pages if (p.status_code or 0) >= 200 and (p.status_code or 0) < 300]

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def style_header(ws, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def add_row(ws, row_num, values):
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # ── Sheet 1: All URLs ──
    ws = wb.active
    ws.title = "All URLs"
    cols = ["URL", "Status Code", "Score", "Title", "Title Length", "Meta Desc Length",
            "H1 Count", "H2 Count", "Word Count", "Text/HTML %", "Internal Links",
            "External Links", "Total Images", "Missing Alt", "Empty Alt", "Response Time (s)",
            "Noindex", "Nofollow", "Has Schema", "Has Hreflang", "Has Viewport", "Issues Count"]
    style_header(ws, cols)
    for i, pg in enumerate(pages, 2):
        add_row(ws, i, [
            pg.url, pg.status_code, pg.score, pg.title, pg.title_length or 0,
            pg.meta_description_length or 0, pg.h1_count or 0, pg.h2_count or 0,
            pg.word_count or 0, pg.code_to_text_ratio,
            pg.internal_links or 0, pg.external_links or 0,
            pg.total_images or 0, pg.images_without_alt or 0, pg.images_with_empty_alt or 0,
            round(pg.response_time, 3) if pg.response_time else None,
            "Yes" if pg.is_noindex else "No", "Yes" if pg.is_nofollow_meta else "No",
            "Yes" if pg.has_schema_markup else "No", "Yes" if pg.has_hreflang else "No",
            "Yes" if pg.has_viewport_meta else "No", len(pg.issues) if pg.issues else 0,
        ])
    ws.column_dimensions["A"].width = 60
    for col_letter in "BCDEFGHIJKLMNOPQRSTUV":
        ws.column_dimensions[col_letter].width = 14

    # ── Issue sheets helper ──
    def add_issue_sheet(title, pages_list, extra_cols=None, extra_fn=None):
        if not pages_list:
            return
        safe_title = title[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=safe_title)
        cols = ["URL"]
        if extra_cols:
            cols.extend(extra_cols)
        style_header(ws, cols)
        for i, pg in enumerate(pages_list, 2):
            values = [pg.url if hasattr(pg, 'url') else pg.get('url', '')]
            if extra_fn:
                values.extend(extra_fn(pg))
            add_row(ws, i, values)
        ws.column_dimensions["A"].width = 60
        for c in range(2, len(cols) + 1):
            ws.column_dimensions[chr(64 + c)].width = 20

    # ── Missing Title ──
    add_issue_sheet("Missing Title", [p for p in content_pages if not p.title])

    # ── Missing Meta Description ──
    add_issue_sheet("Missing Meta Desc", [p for p in content_pages if not p.meta_description])

    # ── Missing H1 ──
    add_issue_sheet("Missing H1", [p for p in content_pages if p.h1_count == 0])

    # ── Missing Viewport ──
    add_issue_sheet("Missing Viewport", [p for p in content_pages if not p.has_viewport_meta])

    # ── Duplicate Titles ──
    title_groups = defaultdict(list)
    for p in content_pages:
        if p.title:
            title_groups[p.title].append(p)
    dup_title_pages = []
    for title_val, pgs in title_groups.items():
        if len(pgs) > 1:
            for pg in pgs:
                dup_title_pages.append({"url": pg.url, "title": title_val, "group_size": len(pgs)})
    if dup_title_pages:
        ws = wb.create_sheet(title="Duplicate Titles")
        style_header(ws, ["URL", "Title", "Duplicated With (count)"])
        for i, d in enumerate(dup_title_pages, 2):
            add_row(ws, i, [d["url"], d["title"], d["group_size"]])
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 20

    # ── Duplicate Meta Descriptions ──
    meta_groups = defaultdict(list)
    for p in content_pages:
        if p.meta_description:
            meta_groups[p.meta_description].append(p)
    dup_meta_pages = []
    for meta_val, pgs in meta_groups.items():
        if len(pgs) > 1:
            for pg in pgs:
                dup_meta_pages.append({"url": pg.url, "meta": meta_val[:100], "group_size": len(pgs)})
    if dup_meta_pages:
        ws = wb.create_sheet(title="Duplicate Meta Desc")
        style_header(ws, ["URL", "Meta Description", "Duplicated With (count)"])
        for i, d in enumerate(dup_meta_pages, 2):
            add_row(ws, i, [d["url"], d["meta"], d["group_size"]])
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 20

    # ── Canonical Issues ──
    canon_pages = [p for p in content_pages if p.canonical_issues and len(p.canonical_issues) > 0]
    if canon_pages:
        ws = wb.create_sheet(title="Canonical Issues")
        style_header(ws, ["URL", "Canonical URL", "Issues"])
        for i, pg in enumerate(canon_pages, 2):
            add_row(ws, i, [pg.url, pg.canonical_url or "none", ", ".join(pg.canonical_issues or [])])
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 40

    # ── Noindex ──
    add_issue_sheet("Noindex Pages", [p for p in content_pages if p.is_noindex])

    # ── Nofollow ──
    add_issue_sheet("Nofollow Pages", [p for p in content_pages if p.is_nofollow_meta])

    # ── Images Missing Alt ──
    img_miss = [p for p in content_pages if p.images_without_alt and p.images_without_alt > 0]
    add_issue_sheet("Images Missing Alt", img_miss,
                    extra_cols=["Missing Count", "Total Images", "Sample Image URL"],
                    extra_fn=lambda pg: [pg.images_without_alt, pg.total_images, (pg.images_without_alt_urls or [None])[0] or ""])

    # ── Images Empty Alt ──
    img_empty = [p for p in content_pages if p.images_with_empty_alt and p.images_with_empty_alt > 0]
    add_issue_sheet("Images Empty Alt", img_empty,
                    extra_cols=["Empty Alt Count", "Total Images"],
                    extra_fn=lambda pg: [pg.images_with_empty_alt, pg.total_images])

    # ── Hreflang Issues ──
    hreflang_pgs = [p for p in content_pages if p.hreflang_issues and len(p.hreflang_issues) > 0]
    add_issue_sheet("Hreflang Issues", hreflang_pgs,
                    extra_cols=["Issues"],
                    extra_fn=lambda pg: ["; ".join(pg.hreflang_issues or [])])

    # ── Thin Content ──
    thin_pgs = [p for p in content_pages if p.word_count and p.word_count < 300]
    add_issue_sheet("Thin Content", thin_pgs,
                    extra_cols=["Word Count"],
                    extra_fn=lambda pg: [pg.word_count or 0])

    # ── Low Text/HTML Ratio ──
    low_ratio = [p for p in content_pages if p.code_to_text_ratio is not None and p.code_to_text_ratio < 10]
    add_issue_sheet("Low Text Ratio", low_ratio,
                    extra_cols=["Text/HTML %"],
                    extra_fn=lambda pg: [pg.code_to_text_ratio])

    # ── Placeholder Content ──
    placeholder = [p for p in content_pages if p.has_placeholders]
    add_issue_sheet("Placeholder Content", placeholder,
                    extra_cols=["Found Text"],
                    extra_fn=lambda pg: [", ".join([(c.get("match") or "")[:30] for c in (pg.placeholder_content or [])[:3]])])

    # ── Slow Pages ──
    slow = [p for p in pages if p.response_time and p.response_time > 3]
    add_issue_sheet("Slow Pages", slow,
                    extra_cols=["Response Time (s)"],
                    extra_fn=lambda pg: [round(pg.response_time, 3) if pg.response_time else 0])

    # ── No Schema ──
    add_issue_sheet("No Schema Markup", [p for p in content_pages if not p.has_schema_markup])

    # ── Redirects ──
    redir = [p for p in pages if (p.status_code or 0) in REDIRECT_CODES]
    add_issue_sheet("Redirects", redir,
                    extra_cols=["Status Code", "Redirect Target"],
                    extra_fn=lambda pg: [pg.status_code, pg.redirect_target or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=seo-report-crawl-{crawl_id}.xlsx"},
    )


# ─── PDF Export ────────────────────────────────────────────
@router.get("/crawls/{crawl_id}/export/pdf")
async def export_crawl_pdf(crawl_id: int, db: AsyncSession = Depends(get_db)):
    """Generate a state-of-the-art PDF report with colored backgrounds, charts, and professional layout."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
        KeepTogether, Flowable
    )
    from reportlab.graphics.shapes import Drawing, String, Wedge, Circle, Rect, Line
    from reportlab.pdfgen import canvas as pdfcanvas
    import math
    from sqlalchemy.orm import selectinload

    crawl_result = await db.execute(
        select(Crawl).options(selectinload(Crawl.project)).where(Crawl.id == crawl_id)
    )
    crawl = crawl_result.scalar_one_or_none()
    if not crawl:
        raise HTTPException(status_code=404, detail="Crawl not found")

    result = await db.execute(select(Page).where(Page.crawl_id == crawl_id))
    pages = result.scalars().all()
    if not pages:
        raise HTTPException(status_code=404, detail="No pages found")

    total = len(pages)
    REDIRECT_CODES = {301, 302, 303, 307, 308}
    content_pages = [p for p in pages if (p.status_code or 0) >= 200 and (p.status_code or 0) < 300]
    content_total = len(content_pages) or 1
    avg_score = round(sum(p.score or 0 for p in content_pages) / content_total, 1)
    critical = sum(1 for p in content_pages for i in (p.issues or []) if i.get("severity") == "critical")
    warnings_count = sum(1 for p in content_pages for i in (p.issues or []) if i.get("severity") == "warning")
    info_count = sum(1 for p in content_pages for i in (p.issues or []) if i.get("severity") == "info")

    # Collect issue data
    missing_title_pages = [p for p in content_pages if not p.title]
    missing_meta_pages = [p for p in content_pages if not p.meta_description]
    missing_h1_pages = [p for p in content_pages if p.h1_count == 0]
    missing_viewport_pages = [p for p in content_pages if not p.has_viewport_meta]
    missing_canonical_pages = [p for p in content_pages if p.canonical_issues and "missing" in p.canonical_issues]
    no_schema_pages = [p for p in content_pages if not p.has_schema_markup]
    noindex_pages = [p for p in content_pages if p.is_noindex]
    nofollow_pages = [p for p in content_pages if p.is_nofollow_meta]
    canon_issues = [p for p in content_pages if p.canonical_issues and len(p.canonical_issues) > 0]
    hreflang_issue_pages = [p for p in content_pages if p.hreflang_issues and len(p.hreflang_issues) > 0]
    img_missing_alt = [p for p in content_pages if p.images_without_alt and p.images_without_alt > 0]
    img_empty_alt = [p for p in content_pages if p.images_with_empty_alt and p.images_with_empty_alt > 0]
    thin_pages = [p for p in content_pages if p.word_count and p.word_count < 300]
    low_ratio_pages = [p for p in content_pages if p.code_to_text_ratio is not None and p.code_to_text_ratio < 10]
    placeholder_pgs = [p for p in content_pages if p.has_placeholders]
    slow_pages = [p for p in pages if p.response_time and p.response_time > 3]
    redirect_pages = [p for p in pages if (p.status_code or 0) in REDIRECT_CODES]
    error_4xx = [p for p in pages if p.status_code and p.status_code >= 400 and p.status_code < 500]
    error_5xx = [p for p in pages if p.status_code and p.status_code >= 500]

    title_groups = defaultdict(list)
    for p in content_pages:
        if p.title:
            title_groups[p.title].append(p)
    dup_titles = {t: pgs for t, pgs in title_groups.items() if len(pgs) > 1}
    meta_groups = defaultdict(list)
    for p in content_pages:
        if p.meta_description:
            meta_groups[p.meta_description].append(p)
    dup_metas = {m: pgs for m, pgs in meta_groups.items() if len(pgs) > 1}

    # Additional issue page lists
    short_title_pages = [p for p in content_pages if p.title_length and 0 < p.title_length < 30]
    long_title_pages = [p for p in content_pages if p.title_length and p.title_length > 60]
    short_meta_pages = [p for p in content_pages if p.meta_description_length and 0 < p.meta_description_length < 120]
    long_meta_pages = [p for p in content_pages if p.meta_description_length and p.meta_description_length > 160]
    multi_h1_pages = [p for p in content_pages if p.h1_count and p.h1_count > 1]
    missing_og_title_pages = [p for p in content_pages if not p.og_title]
    missing_og_image_pages = [p for p in content_pages if not p.og_image]
    no_lazy_pages = [p for p in content_pages if not p.has_lazy_loading]

    sitemaps = crawl.sitemaps_found or []
    sitemap_url_count = sum(sm.get("urls_count", 0) for sm in sitemaps)

    # ── Color palette ──
    PRIMARY = colors.HexColor("#6c5ce7")
    PRIMARY_LIGHT = colors.HexColor("#a29bfe")
    PRIMARY_DARK = colors.HexColor("#4834d4")
    DARK = colors.HexColor("#2d3436")
    DARK2 = colors.HexColor("#636e72")
    WHITE = colors.white
    LIGHT_BG = colors.HexColor("#f5f6fa")
    CARD_BG = colors.HexColor("#fafbfc")
    RED = colors.HexColor("#e74c3c")
    RED_LIGHT = colors.HexColor("#fef0ef")
    ORANGE = colors.HexColor("#e17055")
    ORANGE_LIGHT = colors.HexColor("#fef5f0")
    GREEN = colors.HexColor("#00b894")
    GREEN_LIGHT = colors.HexColor("#edfcf5")
    BLUE = colors.HexColor("#0984e3")
    BLUE_LIGHT = colors.HexColor("#edf5fd")
    PURPLE_LIGHT = colors.HexColor("#f3f0ff")
    YELLOW = colors.HexColor("#fdcb6e")
    GRAY = colors.HexColor("#b2bec3")
    BORDER = colors.HexColor("#dfe6e9")

    W = A4[0]
    page_w = W - 30 * mm  # usable width

    # ── Build PDF ──
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=18 * mm, leftMargin=15 * mm, rightMargin=15 * mm)
    styles = getSampleStyleSheet()

    # Custom styles
    styles.add(ParagraphStyle("CoverTitle", parent=styles["Title"], fontSize=28, leading=34, textColor=WHITE, alignment=TA_CENTER, spaceAfter=6))
    styles.add(ParagraphStyle("CoverSub", parent=styles["Normal"], fontSize=12, leading=16, textColor=colors.HexColor("#dcd6ff"), alignment=TA_CENTER))
    styles.add(ParagraphStyle("CoverDate", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#b8b0e8"), alignment=TA_CENTER))
    styles.add(ParagraphStyle("SectionTitle", parent=styles["Heading1"], fontSize=16, leading=20, textColor=PRIMARY_DARK, spaceBefore=18, spaceAfter=10))
    styles.add(ParagraphStyle("SectionSub", parent=styles["Normal"], fontSize=9, leading=13, textColor=DARK2, spaceAfter=12))
    styles.add(ParagraphStyle("ChapterTitle", parent=styles["Heading2"], fontSize=13, leading=17, textColor=PRIMARY, spaceBefore=10, spaceAfter=6))
    styles.add(ParagraphStyle("ChapterDesc", parent=styles["Normal"], fontSize=8.5, leading=12, textColor=DARK2, spaceAfter=8))
    styles.add(ParagraphStyle("Body9", parent=styles["Normal"], fontSize=9, leading=12, textColor=DARK))
    styles.add(ParagraphStyle("Body8", parent=styles["Normal"], fontSize=8, leading=10, textColor=DARK))
    styles.add(ParagraphStyle("Body7", parent=styles["Normal"], fontSize=7, leading=9, textColor=DARK))
    styles.add(ParagraphStyle("Tiny", parent=styles["Normal"], fontSize=6.5, leading=8, textColor=DARK2))
    styles.add(ParagraphStyle("FooterStyle", parent=styles["Normal"], fontSize=7, textColor=GRAY, alignment=TA_CENTER))
    styles.add(ParagraphStyle("CardValue", parent=styles["Normal"], fontSize=22, leading=26, textColor=PRIMARY_DARK, alignment=TA_CENTER))
    styles.add(ParagraphStyle("CardLabel", parent=styles["Normal"], fontSize=7.5, leading=10, textColor=DARK2, alignment=TA_CENTER))
    styles.add(ParagraphStyle("ScoreGood", parent=styles["Normal"], fontSize=22, leading=26, textColor=GREEN, alignment=TA_CENTER))
    styles.add(ParagraphStyle("ScoreOk", parent=styles["Normal"], fontSize=22, leading=26, textColor=ORANGE, alignment=TA_CENTER))
    styles.add(ParagraphStyle("ScoreBad", parent=styles["Normal"], fontSize=22, leading=26, textColor=RED, alignment=TA_CENTER))
    styles.add(ParagraphStyle("CritVal", parent=styles["Normal"], fontSize=22, leading=26, textColor=RED, alignment=TA_CENTER))
    styles.add(ParagraphStyle("WarnVal", parent=styles["Normal"], fontSize=22, leading=26, textColor=ORANGE, alignment=TA_CENTER))
    styles.add(ParagraphStyle("InfoVal", parent=styles["Normal"], fontSize=22, leading=26, textColor=BLUE, alignment=TA_CENTER))

    def p(text, style="Body7"):
        return Paragraph(str(text), styles[style])

    def url_p(url, max_len=60):
        t = str(url)[:max_len] + ("..." if len(str(url)) > max_len else "")
        return Paragraph(t, styles["Tiny"])

    # ── Table styles ──
    def pro_table_style(accent_color=PRIMARY):
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), accent_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("ALIGN", (0, 0), (-1, 0), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, 0), 1, accent_color),
            ("LINEBELOW", (0, 1), (-1, -2), 0.3, BORDER),
            ("LINEBELOW", (0, -1), (-1, -1), 0.5, accent_color),
            ("ROUNDEDCORNERS", [4, 4, 4, 4]),
        ])

    # ── Donut chart ──
    def make_donut(data_items, width=200, height=140, inner_ratio=0.55, title=""):
        d = Drawing(width, height)
        total_val = sum(v for _, v, _ in data_items if v > 0) or 1
        cx, cy = width * 0.32, height * 0.50
        r_outer = min(width, height) * 0.32
        r_inner = r_outer * inner_ratio
        start = 90
        for label, val, clr in data_items:
            if val <= 0:
                continue
            extent = (val / total_val) * 360
            w_o = Wedge(cx, cy, r_outer, start - extent, start, fillColor=clr, strokeColor=WHITE, strokeWidth=1.5)
            d.add(w_o)
            start -= extent
        # Inner white circle for donut effect
        d.add(Wedge(cx, cy, r_inner, 0, 360, fillColor=WHITE, strokeColor=WHITE, strokeWidth=0))
        # Center text
        if title:
            d.add(String(cx - len(title) * 2, cy - 3, title, fontSize=7, fillColor=DARK2, textAnchor="start"))

        # Legend
        lx = width * 0.68
        ly = height - 16
        for label, val, clr in data_items:
            if val <= 0:
                continue
            pct = round(val / total_val * 100, 1)
            d.add(Rect(lx, ly, 8, 8, fillColor=clr, strokeColor=clr, strokeWidth=0, rx=2, ry=2))
            d.add(String(lx + 12, ly + 1, f"{label}", fontSize=7, fillColor=DARK))
            d.add(String(lx + 12, ly - 8, f"{val} ({pct}%)", fontSize=6.5, fillColor=DARK2))
            ly -= 22
        return d

    # ── Colored background row (for section headers) ──
    class ColoredBlock(Flowable):
        """A colored background block to wrap content visually."""
        def __init__(self, w, h, color, radius=4):
            Flowable.__init__(self)
            self.w = w
            self.h = h
            self.color = color
            self.radius = radius
        def wrap(self, availW, availH):
            return self.w, self.h
        def draw(self):
            self.canv.setFillColor(self.color)
            self.canv.roundRect(0, 0, self.w, self.h, self.radius, fill=1, stroke=0)

    # ── Horizontal bar chart ──
    def make_bar_chart(data_items, width=460, bar_height=16, max_val=None):
        """data_items: list of (label, value, color)"""
        if not data_items:
            return Spacer(1, 1)
        if max_val is None:
            max_val = max(v for _, v, _ in data_items) or 1
        spacing = 6
        total_h = len(data_items) * (bar_height + spacing) + 10
        d = Drawing(width, total_h)
        label_w = 140
        bar_w = width - label_w - 50
        y = total_h - bar_height - 4
        for label, val, clr in data_items:
            if val <= 0:
                y -= bar_height + spacing
                continue
            d.add(String(0, y + 3, label, fontSize=7, fillColor=DARK))
            bw = (val / max_val) * bar_w if max_val else 0
            d.add(Rect(label_w, y, max(bw, 2), bar_height - 2, fillColor=clr, strokeColor=clr, strokeWidth=0, rx=3, ry=3))
            d.add(String(label_w + bw + 4, y + 3, str(val), fontSize=7, fillColor=DARK2))
            y -= bar_height + spacing
        return d

    story = []
    site_url = crawl.project.url if crawl.project else "N/A"
    site_name = crawl.project.name if crawl.project else "N/A"
    report_date = (crawl.completed_at or crawl.created_at or "N/A")
    if hasattr(report_date, "strftime"):
        report_date = report_date.strftime("%B %d, %Y at %H:%M")

    # ════════════════════════════════════════════════════
    # COVER PAGE — purple background
    # ════════════════════════════════════════════════════
    cover_bg = Table([[""]],
        colWidths=[page_w + 4 * mm], rowHeights=[110 * mm])
    cover_bg.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), PRIMARY),
        ("ROUNDEDCORNERS", [8, 8, 8, 8]),
    ]))
    story.append(Spacer(1, 15 * mm))
    story.append(cover_bg)

    # Overlay title (rendered on top via negative spacer)
    story.append(Spacer(1, -95 * mm))
    story.append(Paragraph("SEO Audit Report", styles["CoverTitle"]))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>{site_name}</b>", styles["CoverSub"]))
    story.append(Spacer(1, 2))
    story.append(Paragraph(site_url, styles["CoverSub"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(report_date, styles["CoverDate"]))
    story.append(Spacer(1, 4))
    story.append(Paragraph("Generated by SEO Crawler Pro", styles["CoverDate"]))

    story.append(Spacer(1, 45 * mm))

    # ════════════════════════════════════════════════════
    # PAGE 2 — EXECUTIVE SUMMARY
    # ════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("Executive Summary", styles["SectionTitle"]))

    score_style = "ScoreGood" if avg_score >= 70 else ("ScoreOk" if avg_score >= 40 else "ScoreBad")

    # Scorecard row
    def card_cell(value, label, val_style="CardValue"):
        return Table(
            [[Paragraph(str(value), styles[val_style])], [Paragraph(label, styles["CardLabel"])]],
            colWidths=[page_w / 5 - 4],
            rowHeights=[28, 14]
        )

    cards = Table([[
        card_cell(avg_score, "SEO Score", score_style),
        card_cell(total, "Pages Crawled"),
        card_cell(critical, "Critical", "CritVal"),
        card_cell(warnings_count, "Warnings", "WarnVal"),
        card_cell(info_count, "Info", "InfoVal"),
    ]], colWidths=[page_w / 5] * 5)
    cards.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), CARD_BG),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
        ("LINEBELOW", (0, 0), (-1, -1), 0, WHITE),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
    ]))
    story.append(cards)
    story.append(Spacer(1, 6))

    # Summary sentence
    summary_text = f"We crawled <b>{total}</b> pages on <b>{site_url}</b>. "
    if avg_score >= 70:
        summary_text += "The site is in <b>good shape</b> overall. "
    elif avg_score >= 40:
        summary_text += "The site has <b>several issues</b> that need attention. "
    else:
        summary_text += "The site has <b>significant SEO problems</b> requiring immediate action. "
    summary_text += f"We identified <b>{critical}</b> critical issues, <b>{warnings_count}</b> warnings, and <b>{info_count}</b> informational items."
    story.append(Paragraph(summary_text, styles["SectionSub"]))
    story.append(Spacer(1, 6))

    # ── Charts side by side ──
    story.append(Paragraph("Visual Overview", styles["ChapterTitle"]))

    donut1 = make_donut([
        ("Critical", critical, RED),
        ("Warning", warnings_count, ORANGE),
        ("Info", info_count, BLUE),
    ], width=230, height=130, title="Issues")

    # Status code groups
    sc_groups = defaultdict(int)
    for pg in pages:
        sc = pg.status_code or 0
        if 200 <= sc < 300: sc_groups["2xx"] += 1
        elif 300 <= sc < 400: sc_groups["3xx"] += 1
        elif 400 <= sc < 500: sc_groups["4xx"] += 1
        elif sc >= 500: sc_groups["5xx"] += 1

    donut2 = make_donut([
        ("2xx OK", sc_groups.get("2xx", 0), GREEN),
        ("3xx Redirect", sc_groups.get("3xx", 0), PRIMARY_LIGHT),
        ("4xx Error", sc_groups.get("4xx", 0), ORANGE),
        ("5xx Error", sc_groups.get("5xx", 0), RED),
    ], width=230, height=130, title="Status")

    charts_row = Table([[donut1, donut2]], colWidths=[page_w / 2, page_w / 2])
    charts_row.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
    ]))
    story.append(charts_row)
    story.append(Spacer(1, 10))

    # ── Crawled vs Sitemap mini cards ──
    vs_data = Table([
        [Paragraph("Crawled Pages", styles["CardLabel"]),
         Paragraph("Sitemap URLs", styles["CardLabel"]),
         Paragraph("Avg Response", styles["CardLabel"]),
         Paragraph("Redirects", styles["CardLabel"])],
        [Paragraph(f"<b>{total}</b>", styles["Body9"]),
         Paragraph(f"<b>{sitemap_url_count}</b>", styles["Body9"]),
         Paragraph(f"<b>{round(sum(p.response_time or 0 for p in content_pages) / content_total, 2)}s</b>", styles["Body9"]),
         Paragraph(f"<b>{len(redirect_pages)}</b>", styles["Body9"])],
    ], colWidths=[page_w / 4] * 4)
    vs_data.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, -1), WHITE),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("ROUNDEDCORNERS", [4, 4, 4, 4]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, 0), 0.3, BORDER),
    ]))
    story.append(vs_data)

    # ════════════════════════════════════════════════════
    # PAGE 3 — ISSUE BREAKDOWN
    # ════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("Issue Breakdown", styles["SectionTitle"]))
    story.append(Paragraph("All detected SEO issues ranked by severity. Focus on critical items first.", styles["SectionSub"]))

    issue_rows_data = [
        ("Missing Title Tag", len(missing_title_pages), "Critical", RED),
        ("Missing Meta Description", len(missing_meta_pages), "Critical", RED),
        ("Missing H1 Tag", len(missing_h1_pages), "Critical", RED),
        ("Missing Viewport", len(missing_viewport_pages), "Critical", RED),
        ("Placeholder Content", len(placeholder_pgs), "Critical", RED),
        ("4xx Client Errors", len(error_4xx), "Critical", RED),
        ("5xx Server Errors", len(error_5xx), "Critical", RED),
        ("Missing Canonical", len(missing_canonical_pages), "Warning", ORANGE),
        ("Canonical Issues", len(canon_issues), "Warning", ORANGE),
        ("Duplicate Titles", len(dup_titles), "Warning", ORANGE),
        ("Duplicate Metas", len(dup_metas), "Warning", ORANGE),
        ("Short Title (<30 chars)", len(short_title_pages), "Warning", ORANGE),
        ("Long Title (>60 chars)", len(long_title_pages), "Warning", ORANGE),
        ("Short Meta Desc (<120)", len(short_meta_pages), "Warning", ORANGE),
        ("Long Meta Desc (>160)", len(long_meta_pages), "Warning", ORANGE),
        ("Multiple H1 Tags", len(multi_h1_pages), "Warning", ORANGE),
        ("Noindex Pages", len(noindex_pages), "Warning", ORANGE),
        ("Nofollow Pages", len(nofollow_pages), "Warning", ORANGE),
        ("Images Missing Alt", len(img_missing_alt), "Warning", ORANGE),
        ("Images Empty Alt", len(img_empty_alt), "Warning", ORANGE),
        ("Thin Content (<300 words)", len(thin_pages), "Warning", ORANGE),
        ("Low Text/HTML Ratio", len(low_ratio_pages), "Warning", ORANGE),
        ("Slow Pages (>3s)", len(slow_pages), "Warning", ORANGE),
        ("Hreflang Issues", len(hreflang_issue_pages), "Warning", ORANGE),
        ("Redirects", len(redirect_pages), "Info", PRIMARY_LIGHT),
        ("No Schema Markup", len(no_schema_pages), "Info", BLUE),
        ("Missing OG Title", len(missing_og_title_pages), "Info", BLUE),
        ("Missing OG Image", len(missing_og_image_pages), "Info", BLUE),
        ("No Lazy Loading", len(no_lazy_pages), "Info", BLUE),
    ]
    # Filter to only non-zero
    active_issues = [(n, c, s, cl) for n, c, s, cl in issue_rows_data if c > 0]

    if active_issues:
        # Bar chart
        bar_data = [(n, c, cl) for n, c, s, cl in active_issues[:12]]
        story.append(make_bar_chart(bar_data, width=page_w))
        story.append(Spacer(1, 10))

        # Table
        tbl_data = [["Issue", "Count", "Severity"]]
        for name, count, sev, clr in active_issues:
            sev_color = "#e74c3c" if sev == "Critical" else ("#e17055" if sev == "Warning" else "#0984e3")
            tbl_data.append([
                Paragraph(name, styles["Body8"]),
                Paragraph(f"<b>{count}</b>", styles["Body8"]),
                Paragraph(f'<font color="{sev_color}"><b>{sev}</b></font>', styles["Body8"]),
            ])
        t = Table(tbl_data, colWidths=[page_w * 0.55, page_w * 0.2, page_w * 0.25])
        t.setStyle(pro_table_style(PRIMARY))
        story.append(t)

    # ════════════════════════════════════════════════════
    # ISSUE DETAIL CHAPTERS (3 URL examples each)
    # ════════════════════════════════════════════════════
    ch_num = 1

    def issue_chapter(title, description, pages_list, cols, row_fn, severity_color=PRIMARY, max_rows=5):
        nonlocal ch_num
        if not pages_list:
            return

        story.append(Spacer(1, 14))

        # Colored severity bar
        sev_bar = Table(
            [[Paragraph(f"<b>{ch_num}. {title}</b>", styles["Body8"]),
              Paragraph(f"<b>{len(pages_list)} pages affected</b>", styles["Body8"])]],
            colWidths=[page_w * 0.6, page_w * 0.4]
        )
        sev_bar.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), severity_color),
            ("TEXTCOLOR", (0, 0), (-1, -1), WHITE),
            ("ROUNDEDCORNERS", [6, 6, 0, 0]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("RIGHTPADDING", (1, 0), (1, 0), 10),
        ]))
        story.append(sev_bar)

        # Paragraph description (flows naturally, no big whitespace)
        story.append(Spacer(1, 4))
        story.append(Paragraph(description, styles["ChapterDesc"]))
        ch_num += 1

        d = [cols]
        for pg in pages_list[:max_rows]:
            d.append(row_fn(pg))
        if len(pages_list) > max_rows:
            remaining = len(pages_list) - max_rows
            d.append([Paragraph(f"... and {remaining} more URLs. See Excel export for complete list.", styles["Tiny"])] + [""] * (len(cols) - 1))
        t = Table(d, colWidths=[int(page_w / len(cols))] * len(cols))
        t.setStyle(pro_table_style(severity_color))
        story.append(KeepTogether([t]))

    # ── Critical issues ──
    issue_chapter("Missing Title Tag",
                  "Every page needs a unique, descriptive title tag. Search engines display this in results and use it as a primary ranking signal.",
                  missing_title_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], RED)

    issue_chapter("Missing Meta Description",
                  "Meta descriptions appear in search results below the title. A compelling description improves click-through rates.",
                  missing_meta_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], RED)

    issue_chapter("Missing H1 Tag",
                  "The H1 tag defines the main topic of the page. Every page should have exactly one H1 heading.",
                  missing_h1_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], RED)

    issue_chapter("Missing Viewport Meta",
                  "Without a viewport meta tag, mobile devices won't render the page correctly. This directly impacts mobile rankings.",
                  missing_viewport_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], RED)

    issue_chapter("Placeholder / Lorem Ipsum Content",
                  "Pages with placeholder text are unfinished and harm user experience and SEO.",
                  placeholder_pgs, ["URL", "Detected Text"],
                  lambda pg: [url_p(pg.url, 40), p(", ".join([(c.get("match") or "")[:30] for c in (pg.placeholder_content or [])[:2]]))], RED)

    issue_chapter("4xx Client Errors",
                  "Pages returning 4xx status codes (404 Not Found, 403 Forbidden, etc.) hurt user experience and waste crawl budget.",
                  error_4xx, ["URL", "Status Code"],
                  lambda pg: [url_p(pg.url, 50), p(str(pg.status_code))], RED)

    issue_chapter("5xx Server Errors",
                  "Server errors indicate infrastructure problems that prevent pages from loading.",
                  error_5xx, ["URL", "Status Code"],
                  lambda pg: [url_p(pg.url, 50), p(str(pg.status_code))], RED)

    # ── Warning issues ──
    issue_chapter("Missing Canonical Tag",
                  "Pages without a canonical tag risk duplicate content issues. Every indexable page should declare its canonical URL.",
                  missing_canonical_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], ORANGE)

    issue_chapter("Canonical Tag Issues",
                  "Incorrect canonical tags send conflicting signals to search engines about which version of a page to index.",
                  [p_item for p_item in canon_issues if p_item.canonical_issues and "missing" not in p_item.canonical_issues],
                  ["URL", "Canonical URL", "Issues"],
                  lambda pg: [url_p(pg.url, 35), url_p(pg.canonical_url or "none", 30), p(", ".join(pg.canonical_issues or [])[:50])], ORANGE)

    # Duplicate Titles
    if dup_titles:
        story.append(Spacer(1, 14))
        sev_bar = Table(
            [[Paragraph(f"<b>{ch_num}. Duplicate Title Tags</b>", styles["Body8"]),
              Paragraph(f"<b>{len(dup_titles)} groups</b>", styles["Body8"])]],
            colWidths=[page_w * 0.6, page_w * 0.4])
        sev_bar.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), ORANGE),
            ("TEXTCOLOR", (0, 0), (-1, -1), WHITE),
            ("ROUNDEDCORNERS", [6, 6, 0, 0]),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"), ("RIGHTPADDING", (1, 0), (1, 0), 10),
        ]))
        story.append(sev_bar)
        story.append(Spacer(1, 4))
        story.append(Paragraph("Multiple pages sharing the same title confuse search engines and dilute ranking potential. Each page should have a unique, descriptive title that accurately represents its content.", styles["ChapterDesc"]))
        ch_num += 1
        shown = 0
        for title_val, pgs in dup_titles.items():
            if shown >= 3:
                break
            story.append(Paragraph(f'<font color="#e17055"><b>"{title_val[:65]}{"..." if len(title_val) > 65 else ""}"</b></font> ({len(pgs)} pages)', styles["Body8"]))
            for pg in pgs[:3]:
                story.append(Paragraph(f"  {pg.url[:75]}", styles["Tiny"]))
            story.append(Spacer(1, 4))
            shown += 1
        if len(dup_titles) > 3:
            story.append(Paragraph(f"... and {len(dup_titles) - 3} more groups. See Excel export.", styles["Tiny"]))

    # Duplicate Meta Descriptions
    if dup_metas:
        story.append(Spacer(1, 14))
        sev_bar = Table(
            [[Paragraph(f"<b>{ch_num}. Duplicate Meta Descriptions</b>", styles["Body8"]),
              Paragraph(f"<b>{len(dup_metas)} groups</b>", styles["Body8"])]],
            colWidths=[page_w * 0.6, page_w * 0.4])
        sev_bar.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), ORANGE),
            ("TEXTCOLOR", (0, 0), (-1, -1), WHITE),
            ("ROUNDEDCORNERS", [6, 6, 0, 0]),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"), ("RIGHTPADDING", (1, 0), (1, 0), 10),
        ]))
        story.append(sev_bar)
        story.append(Spacer(1, 4))
        story.append(Paragraph("Unique meta descriptions for each page improve click-through rates from search results. When multiple pages share the same description, search engines may choose to show a generic snippet instead.", styles["ChapterDesc"]))
        ch_num += 1
        shown = 0
        for meta_val, pgs in dup_metas.items():
            if shown >= 3:
                break
            story.append(Paragraph(f'<font color="#e17055"><b>"{meta_val[:65]}{"..." if len(meta_val) > 65 else ""}"</b></font> ({len(pgs)} pages)', styles["Body8"]))
            for pg in pgs[:3]:
                story.append(Paragraph(f"  {pg.url[:75]}", styles["Tiny"]))
            story.append(Spacer(1, 4))
            shown += 1
        if len(dup_metas) > 3:
            story.append(Paragraph(f"... and {len(dup_metas) - 3} more groups. See Excel export.", styles["Tiny"]))

    issue_chapter("Noindex Pages",
                  "These pages tell search engines not to include them in search results. Verify this is intentional.",
                  noindex_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], ORANGE)

    issue_chapter("Nofollow Meta Pages",
                  "The nofollow meta tag prevents search engines from following links on these pages, blocking link equity flow.",
                  nofollow_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], ORANGE)

    issue_chapter("Images Missing Alt Attribute",
                  "Alt text is essential for accessibility (screen readers) and helps search engines understand image content.",
                  img_missing_alt, ["URL", "Missing", "Total"],
                  lambda pg: [url_p(pg.url, 40), p(str(pg.images_without_alt)), p(str(pg.total_images))], ORANGE)

    issue_chapter("Images With Empty Alt",
                  "Empty alt attributes provide no context. Decorative images should use alt=\"\" but content images need descriptive text.",
                  img_empty_alt, ["URL", "Empty", "Total"],
                  lambda pg: [url_p(pg.url, 40), p(str(pg.images_with_empty_alt or 0)), p(str(pg.total_images))], ORANGE)

    issue_chapter("Hreflang Issues",
                  "Hreflang tags tell search engines which language/region a page targets. Misconfigurations hurt international SEO.",
                  hreflang_issue_pages, ["URL", "Issues Found"],
                  lambda pg: [url_p(pg.url, 40), p("; ".join(pg.hreflang_issues or [])[:70])], ORANGE)

    issue_chapter("Thin Content (under 300 words)",
                  "Pages with very little text content provide limited value to users and typically rank poorly.",
                  thin_pages, ["URL", "Word Count"],
                  lambda pg: [url_p(pg.url, 50), p(str(pg.word_count or 0))], ORANGE)

    issue_chapter("Low Text-to-HTML Ratio (under 10%)",
                  "A low ratio suggests pages are heavy on code and light on readable content.",
                  low_ratio_pages, ["URL", "Ratio"],
                  lambda pg: [url_p(pg.url, 50), p(f"{pg.code_to_text_ratio}%")], ORANGE)

    issue_chapter("Slow Pages (over 3s response)",
                  "Page speed is a confirmed ranking factor. Pages loading over 3 seconds have higher bounce rates.",
                  slow_pages, ["URL", "Response Time"],
                  lambda pg: [url_p(pg.url, 50), p(f"{pg.response_time:.2f}s")], ORANGE)

    issue_chapter("Short Title Tags (under 30 chars)",
                  "Titles under 30 characters may not provide enough context for search engines or users. Aim for 30-60 characters.",
                  short_title_pages, ["URL", "Title", "Length"],
                  lambda pg: [url_p(pg.url, 35), p((pg.title or "")[:40]), p(str(pg.title_length or 0))], ORANGE)

    issue_chapter("Long Title Tags (over 60 chars)",
                  "Titles over 60 characters get truncated in search results, potentially cutting off important information.",
                  long_title_pages, ["URL", "Title", "Length"],
                  lambda pg: [url_p(pg.url, 35), p((pg.title or "")[:40] + "..."), p(str(pg.title_length or 0))], ORANGE)

    issue_chapter("Short Meta Descriptions (under 120 chars)",
                  "Short meta descriptions miss the opportunity to fully describe the page content and attract clicks.",
                  short_meta_pages, ["URL", "Length"],
                  lambda pg: [url_p(pg.url, 50), p(str(pg.meta_description_length or 0))], ORANGE)

    issue_chapter("Long Meta Descriptions (over 160 chars)",
                  "Meta descriptions over 160 characters get truncated in search results.",
                  long_meta_pages, ["URL", "Length"],
                  lambda pg: [url_p(pg.url, 50), p(str(pg.meta_description_length or 0))], ORANGE)

    issue_chapter("Multiple H1 Tags",
                  "Each page should have exactly one H1 tag. Multiple H1 tags dilute the topical focus and confuse search engines.",
                  multi_h1_pages, ["URL", "H1 Count"],
                  lambda pg: [url_p(pg.url, 50), p(str(pg.h1_count or 0))], ORANGE)

    # ── Info issues ──
    issue_chapter("No Schema Markup",
                  "Schema markup (structured data) enables rich snippets in search results, improving visibility and click-through rates.",
                  no_schema_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], BLUE)

    issue_chapter("Missing OG Title",
                  "Open Graph title tags control how pages appear when shared on social media. Missing OG titles may result in poor social previews.",
                  missing_og_title_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], BLUE)

    issue_chapter("Missing OG Image",
                  "Pages without an Open Graph image tag will have no image preview when shared on social media, significantly reducing engagement.",
                  missing_og_image_pages, ["URL"],
                  lambda pg: [url_p(pg.url)], BLUE)

    issue_chapter("Redirects",
                  "Pages returning redirect status codes. Excessive redirects slow page loading and waste crawl budget.",
                  redirect_pages, ["URL", "Status Code"],
                  lambda pg: [url_p(pg.url, 50), p(str(pg.status_code))], PRIMARY_LIGHT)

    # ════════════════════════════════════════════════════
    # ROBOTS & SITEMAPS
    # ════════════════════════════════════════════════════
    story.append(Spacer(1, 18))
    sev_bar = Table(
        [[Paragraph(f"<b>Chapter {ch_num}</b>", styles["Body8"]),
          Paragraph("<b>Technical</b>", styles["Body8"])]],
        colWidths=[page_w * 0.5, page_w * 0.5])
    sev_bar.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), DARK),
        ("TEXTCOLOR", (0, 0), (-1, -1), WHITE),
        ("ROUNDEDCORNERS", [6, 6, 0, 0]),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"), ("RIGHTPADDING", (1, 0), (1, 0), 10),
    ]))
    story.append(sev_bar)
    story.append(Paragraph("Robots.txt and Sitemaps", styles["ChapterTitle"]))

    robots_status = crawl.robots_txt_status or "unknown"
    robots_color = "#00b894" if robots_status == "found" else "#e74c3c"
    story.append(Paragraph(f'robots.txt: <font color="{robots_color}"><b>{robots_status.upper()}</b></font>. '
                           f'{len(sitemaps)} sitemap(s) discovered containing {sitemap_url_count} URLs.', styles["ChapterDesc"]))

    if sitemaps:
        sm_data = [["Sitemap URL", "Type", "URLs"]]
        for sm in sitemaps:
            sm_data.append([url_p(sm.get("url", "?"), 50), p(sm.get("type", "?")), p(str(sm.get("urls_count", 0)))])
        t = Table(sm_data, colWidths=[page_w * 0.60, page_w * 0.22, page_w * 0.18])
        t.setStyle(pro_table_style(DARK))
        story.append(t)

    # ════════════════════════════════════════════════════
    # FOOTER
    # ════════════════════════════════════════════════════
    story.append(Spacer(1, 20))

    # Footer bar
    footer_bar = Table(
        [[Paragraph("Generated by <b>SEO Crawler Pro</b>  |  ai.tudordaniel.ro", styles["FooterStyle"])]],
        colWidths=[page_w])
    footer_bar.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
        ("ROUNDEDCORNERS", [4, 4, 4, 4]),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.3, BORDER),
    ]))
    story.append(footer_bar)

    doc.build(story)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=seo-report-crawl-{crawl_id}.pdf"},
    )
